#!/usr/bin/env python3
"""
Generar planilla Excel del fixture - Torneo 45
Formato: Canchas en columnas, horarios en filas
"""
import sys, os
sys.path.insert(0, os.path.dirname(__file__))
from dotenv import load_dotenv
from sqlalchemy import create_engine, text
from sqlalchemy.orm import sessionmaker
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from collections import defaultdict

load_dotenv()
engine = create_engine(os.getenv("DATABASE_URL"))
Session = sessionmaker(bind=engine)

TORNEO_ID = 45

def generar_excel():
    s = Session()
    try:
        print("=" * 80)
        print(f"GENERAR PLANILLA FIXTURE - TORNEO {TORNEO_ID}")
        print("=" * 80)

        # Obtener información del torneo
        torneo = s.execute(text("SELECT nombre FROM torneos WHERE id = :t"), {"t": TORNEO_ID}).fetchone()
        if not torneo:
            print(f"\n❌ Torneo {TORNEO_ID} no encontrado")
            return
        
        print(f"\n✅ Torneo: {torneo[0]}")

        # Obtener todas las canchas
        canchas = s.execute(text("""
            SELECT id, nombre FROM torneo_canchas 
            WHERE torneo_id = :t AND activa = true
            ORDER BY nombre
        """), {"t": TORNEO_ID}).fetchall()
        
        if not canchas:
            print("\n❌ No hay canchas configuradas")
            return
        
        print(f"✅ Canchas: {len(canchas)}")

        # Obtener todos los partidos con información de zona
        partidos = s.execute(text("""
            SELECT 
                p.id_partido,
                p.fecha_hora,
                p.cancha_id,
                tc.nombre as categoria,
                tz.nombre as zona,
                tz.id as zona_id,
                tp1.id as pareja1_id,
                tp2.id as pareja2_id
            FROM partidos p
            JOIN torneos_parejas tp1 ON p.pareja1_id = tp1.id
            JOIN torneos_parejas tp2 ON p.pareja2_id = tp2.id
            LEFT JOIN torneo_categorias tc ON tp1.categoria_id = tc.id
            LEFT JOIN torneo_zonas tz ON p.zona_id = tz.id
            WHERE tp1.torneo_id = :t
            AND p.fecha_hora IS NOT NULL
            ORDER BY p.fecha_hora
        """), {"t": TORNEO_ID}).fetchall()

        if not partidos:
            print("\n❌ No hay partidos programados")
            return
        
        print(f"✅ Partidos: {len(partidos)}")

        # Obtener posiciones de parejas en zonas
        print("\n📊 Obteniendo posiciones en zonas...")
        posiciones_zona = {}
        zonas_ids = set(p.zona_id for p in partidos if p.zona_id)
        
        for zona_id in zonas_ids:
            parejas_zona = s.execute(text("""
                SELECT pareja_id FROM torneo_zona_parejas
                WHERE zona_id = :zid
                ORDER BY id
            """), {"zid": zona_id}).fetchall()
            
            for idx, (pareja_id,) in enumerate(parejas_zona, 1):
                posiciones_zona[pareja_id] = idx

        # Crear workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Fixture"

        # Estilos
        header_fill = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Header: Título
        ws.merge_cells('A1:' + chr(65 + len(canchas)) + '1')
        ws['A1'] = torneo[0]
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = center_align
        ws['A1'].fill = PatternFill(start_color="E8F4F8", end_color="E8F4F8", fill_type="solid")

        # Header: Columnas (Horario + Canchas)
        ws['A2'] = 'HORARIO'
        ws['A2'].font = header_font
        ws['A2'].fill = header_fill
        ws['A2'].alignment = center_align
        ws['A2'].border = border

        for idx, cancha in enumerate(canchas, 1):
            col_letter = chr(65 + idx)  # B, C, D...
            ws[f'{col_letter}2'] = cancha.nombre
            ws[f'{col_letter}2'].font = header_font
            ws[f'{col_letter}2'].fill = header_fill
            ws[f'{col_letter}2'].alignment = center_align
            ws[f'{col_letter}2'].border = border

        # Agrupar partidos por fecha/hora y cancha
        partidos_por_slot = defaultdict(dict)
        horarios_unicos = set()
        
        for p in partidos:
            fecha_hora = p.fecha_hora
            hora_str = fecha_hora.strftime("%d/%m %H:%M")
            horarios_unicos.add((fecha_hora, hora_str))
            
            cancha_id = p.cancha_id
            
            # Obtener posiciones en zona
            pos1 = posiciones_zona.get(p.pareja1_id, "?")
            pos2 = posiciones_zona.get(p.pareja2_id, "?")
            
            # Formato: "CAT - ZONA\n1 vs 3"
            categoria = p.categoria or "?"
            zona = p.zona or "?"
            texto = f"{categoria} - {zona}\n{pos1} vs {pos2}"
            
            partidos_por_slot[hora_str][cancha_id] = texto

        # Ordenar horarios
        horarios_ordenados = sorted(horarios_unicos, key=lambda x: x[0])

        # Llenar filas
        row = 3
        for fecha_hora, hora_str in horarios_ordenados:
            # Columna de horario
            ws[f'A{row}'] = hora_str
            ws[f'A{row}'].font = Font(bold=True, size=10)
            ws[f'A{row}'].alignment = center_align
            ws[f'A{row}'].border = border
            ws[f'A{row}'].fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")

            # Columnas de canchas
            for idx, cancha in enumerate(canchas, 1):
                col_letter = chr(65 + idx)
                cell = ws[f'{col_letter}{row}']
                
                # Buscar partido en este slot
                if cancha.id in partidos_por_slot[hora_str]:
                    cell.value = partidos_por_slot[hora_str][cancha.id]
                    cell.fill = PatternFill(start_color="E8F4F8", end_color="E8F4F8", fill_type="solid")
                else:
                    cell.value = ""
                
                cell.alignment = center_align
                cell.border = border
                cell.font = Font(size=9)

            row += 1

        # Ajustar anchos de columna
        ws.column_dimensions['A'].width = 15
        for idx in range(len(canchas)):
            col_letter = chr(66 + idx)  # B, C, D...
            ws.column_dimensions[col_letter].width = 20

        # Ajustar alturas de fila
        for r in range(3, row):
            ws.row_dimensions[r].height = 35

        # Guardar archivo
        filename = f"fixture_torneo_{TORNEO_ID}.xlsx"
        wb.save(filename)

        print(f"\n{'=' * 80}")
        print(f"✅ LISTO - Archivo generado: {filename}")
        print(f"   📊 {len(horarios_ordenados)} horarios")
        print(f"   🏟️  {len(canchas)} canchas")
        print(f"   🏐 {len(partidos)} partidos")
        print("=" * 80)

    except Exception as e:
        print(f"\n❌ Error: {e}")
        import traceback
        traceback.print_exc()
    finally:
        s.close()

if __name__ == "__main__":
    generar_excel()
